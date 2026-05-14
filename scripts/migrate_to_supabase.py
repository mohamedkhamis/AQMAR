# scripts/migrate_to_supabase.py
"""One-shot push of data/martyrs.xlsx + data/photos/ → Supabase.

Run after:
  1. Filling .env with SUPABASE_* keys
  2. Running scripts/setup_supabase_schema.sql in the Dashboard
  3. Creating the aqmar-photos bucket (Public)

Idempotent — re-running will UPSERT rows and re-upload photos
(overwrites existing). Safe to re-run after partial failures.
"""
import os
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))

from openpyxl import load_workbook
from src.config import load_config
from src.supabase_client import make_sync_from_config

EXCEL_PATH = "data/martyrs.xlsx"
PHOTOS_DIR = "data/photos"
CHUNK = 50

# Excel columns (1-indexed) per scripts/excel_to_json.py
COLUMNS = [
    ("msg_id",            1,  int),
    ("name",              2,  str),
    ("name_normalized",   3,  str),
    ("birth_date",        4,  str),
    ("martyrdom_date",    5,  str),
    ("city",              6,  str),
    ("military_rank",     7,  str),
    ("weapon",            8,  str),
    ("battalion",         9,  str),
    ("brigade",          10,  str),
    ("photo_path",       11,  str),   # local path, will be rewritten
    ("posted_date",      13,  str),
    ("message_link",     14,  str),
    ("extraction_status",15,  str),
    ("duplicate_status", 16,  str),
]
DUPLICATE_COLUMNS = [
    ("msg_id",        1,  int),
    ("name",          2,  str),
    ("reason",        3,  str),
    ("resolution",    4,  str),
    ("size_mb",       5,  float),
    ("kept_msg_id",   6,  int),
    ("link",          7,  str),
]


def cell_to(value, caster):
    if value is None or value == "":
        return None if caster is str else 0
    try:
        return caster(value)
    except (TypeError, ValueError):
        return None if caster is str else 0


def build_main_rows(ws):
    rows = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if row[0] is None:
            continue
        record = {}
        for key, col, caster in COLUMNS:
            v = row[col - 1] if col - 1 < len(row) else None
            record[key] = cell_to(v, caster)
        rows.append(record)
    return rows


def build_duplicate_rows(ws):
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        record = {}
        for key, col, caster in DUPLICATE_COLUMNS:
            v = row[col - 1] if col - 1 < len(row) else None
            record[key] = cell_to(v, caster)
        rows.append(record)
    return rows


def main():
    cfg = load_config()
    sync = make_sync_from_config(cfg)
    wb = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    main_rows = build_main_rows(wb["الشهداء"])
    dup_rows = build_duplicate_rows(wb["النسخ_المكررة"]) if "النسخ_المكررة" in wb.sheetnames else []

    print(f"Excel: {len(main_rows)} martyrs + {len(dup_rows)} duplicates")
    print(f"Photos dir: {PHOTOS_DIR}")

    # Step 1: Upload photos
    photo_count = 0
    photo_fail = 0
    for r in main_rows:
        msg_id = r["msg_id"]
        local = os.path.join(PHOTOS_DIR, f"{msg_id}.jpg")
        if os.path.exists(local) and os.path.getsize(local) > 0:
            try:
                sync.upload_photo(msg_id, local)
                r["photo_path"] = sync.public_photo_url(msg_id)
                photo_count += 1
                if photo_count % 50 == 0:
                    print(f"  uploaded {photo_count} photos...")
            except Exception as e:
                print(f"  photo upload failed for msg {msg_id}: {e}")
                r["photo_path"] = None
                photo_fail += 1
        else:
            r["photo_path"] = None
    print(f"Photos: {photo_count} uploaded, {photo_fail} failed")

    # Step 2: Upsert martyr rows in chunks
    for i in range(0, len(main_rows), CHUNK):
        chunk = main_rows[i:i + CHUNK]
        try:
            sync.client.table("martyrs").upsert(chunk).execute()
            print(f"  upserted martyrs {i + 1}..{i + len(chunk)} / {len(main_rows)}")
        except Exception as e:
            print(f"  ERROR upserting chunk {i}: {e}")

    # Step 3: Upsert duplicates
    if dup_rows:
        for i in range(0, len(dup_rows), CHUNK):
            chunk = dup_rows[i:i + CHUNK]
            try:
                sync.client.table("martyrs_duplicates").upsert(chunk).execute()
                print(f"  upserted dupes {i + 1}..{i + len(chunk)} / {len(dup_rows)}")
            except Exception as e:
                print(f"  ERROR upserting dupes chunk {i}: {e}")

    print(f"\nDone. Migrated {len(main_rows)} martyrs ({photo_count} photos) + {len(dup_rows)} duplicates.")


if __name__ == "__main__":
    main()
