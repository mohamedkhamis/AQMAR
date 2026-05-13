# scripts/fill_birth_dates_v2.py
"""Faster version of fill_birth_dates.py.

Two improvements vs v1:
  1. 3 concurrent video downloads via asyncio.Semaphore (vs sequential)
  2. 5 frame timestamps (25/28/30/32/35) vs the default 3 (28/30/32) —
     wider net catches videos where the data slide appears at sec 26
     or 34 instead of exactly 30. Same download cost, more OCR per
     video (~1s extra) for much higher hit rate.

Expected runtime for ~70 missing-birth rows: ~2-4 hours
(vs ~10-15h for the sequential v1).

Run from project root:
  python scripts\\fill_birth_dates_v2.py
"""
import asyncio
import logging
import os
import sys
import tempfile
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))
from openpyxl import load_workbook
from src.config import load_config
from src.telegram_client import TelegramFetcher
from src.parser_ocr import parse_ocr_text, merge_extractions
from src.frame_extractor import extract_frames
from src.ocr_engine import ocr_image
from src.state import State

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(message)s",
    handlers=[
        logging.FileHandler("logs/fill_birth_v2.log", encoding="utf-8"),
        logging.StreamHandler(),
    ],
)

EXCEL_PATH = "data/martyrs.xlsx"
STATE_PATH = "data/state.json"
FRAMES_DIR = "data/frames"

# Excel column numbers (1-indexed)
COL_MSG_ID = 1
COL_BIRTH = 4
COL_MARTYRDOM = 5
COL_CITY = 6
COL_RANK = 7
COL_WEAPON = 8
COL_FRAMES = 12
COL_STATUS = 15

FRAME_TIMESTAMPS = [25, 28, 30, 32, 35]
CONCURRENCY = 3
CHECKPOINT_EVERY = 3


def determine_status(birth, martyrdom):
    if not birth and not martyrdom:
        return "missing_critical"
    if not birth:
        return "partial_birth"
    if not martyrdom:
        return "partial_martyrdom"
    return "complete"


async def process_one(fetcher, msg_id, excel_row, ws, state, lock, counter):
    try:
        msg = await fetcher.client.get_messages(fetcher.channel, ids=msg_id)
        if msg is None:
            logging.warning(f"msg {msg_id} not found")
            async with lock:
                counter["fail"] += 1
            return
        tg = fetcher._to_tg_message(msg)
        if not tg.has_video:
            logging.warning(f"msg {msg_id} has no video")
            async with lock:
                counter["fail"] += 1
            return

        # Download video + extract 5 frames
        with tempfile.TemporaryDirectory() as td:
            video_path = os.path.join(td, f"{msg_id}.mp4")
            await fetcher.download_video(tg, video_path)
            frame_paths = extract_frames(video_path, FRAMES_DIR, msg_id, timestamps=FRAME_TIMESTAMPS)

        if not frame_paths:
            logging.warning(f"msg {msg_id}: no frames extracted")
            async with lock:
                counter["fail"] += 1
            return

        # OCR each frame, merge by majority vote
        extractions = []
        for fp in frame_paths:
            try:
                text = ocr_image(fp)
                extractions.append(parse_ocr_text(text))
            except Exception as e:
                logging.warning(f"OCR failed on {fp}: {e}")

        if not extractions:
            async with lock:
                counter["fail"] += 1
            return

        merged = merge_extractions(extractions)

        # Update Excel row (under lock; openpyxl is not async-safe).
        # Never clear an existing value — only fill in blanks.
        async with lock:
            if merged.get("birth_date"):
                ws.cell(row=excel_row, column=COL_BIRTH, value=merged["birth_date"])
            if merged.get("martyrdom_date") and not ws.cell(row=excel_row, column=COL_MARTYRDOM).value:
                ws.cell(row=excel_row, column=COL_MARTYRDOM, value=merged["martyrdom_date"])
            if merged.get("city") and not ws.cell(row=excel_row, column=COL_CITY).value:
                ws.cell(row=excel_row, column=COL_CITY, value=merged["city"])
            if merged.get("military_rank") and not ws.cell(row=excel_row, column=COL_RANK).value:
                ws.cell(row=excel_row, column=COL_RANK, value=merged["military_rank"])
            if merged.get("weapon") and not ws.cell(row=excel_row, column=COL_WEAPON).value:
                ws.cell(row=excel_row, column=COL_WEAPON, value=merged["weapon"])
            ws.cell(row=excel_row, column=COL_FRAMES, value=";".join(frame_paths))

            birth_now = ws.cell(row=excel_row, column=COL_BIRTH).value
            mart_now = ws.cell(row=excel_row, column=COL_MARTYRDOM).value
            new_status = determine_status(birth_now, mart_now)
            ws.cell(row=excel_row, column=COL_STATUS, value=new_status)
            state.mark_processed(msg_id, new_status)
            counter["done"] += 1
            print(f"  [{counter['done']}/{counter['total']}] msg {msg_id}: {new_status} | "
                  f"birth={birth_now} | mart={mart_now}")
    except Exception as e:
        logging.exception(f"Failed msg {msg_id}: {e}")
        async with lock:
            state.mark_processed(msg_id, "video_failed")
            counter["fail"] += 1


async def main():
    cfg = load_config()
    fetcher = TelegramFetcher(
        cfg.api_id, cfg.api_hash, cfg.phone, cfg.two_fa_password,
        cfg.session_path, cfg.channel_username,
    )
    await fetcher.connect()

    wb = load_workbook(EXCEL_PATH)
    ws = wb["الشهداء"]
    state = State.load(STATE_PATH)

    # Find every Excel row that still has an empty birth_date cell.
    targets = []
    for r in range(3, ws.max_row + 1):
        birth = ws.cell(row=r, column=COL_BIRTH).value
        msg_id = ws.cell(row=r, column=COL_MSG_ID).value
        if msg_id and not birth:
            targets.append((r, int(msg_id)))
    print(f"Found {len(targets)} rows needing birth date.")
    print(f"Concurrency: {CONCURRENCY} parallel downloads")
    print(f"Frame timestamps: {FRAME_TIMESTAMPS}")
    print()
    if not targets:
        await fetcher.disconnect()
        return

    counter = {"done": 0, "fail": 0, "total": len(targets)}
    sem = asyncio.Semaphore(CONCURRENCY)
    lock = asyncio.Lock()

    async def run_one(excel_row, msg_id):
        async with sem:
            await process_one(fetcher, msg_id, excel_row, ws, state, lock, counter)
            if (counter["done"] + counter["fail"]) % CHECKPOINT_EVERY == 0:
                async with lock:
                    wb.save(EXCEL_PATH)
                    state.save(STATE_PATH)
                    print(f"  ✓ checkpoint at {counter['done'] + counter['fail']}/{counter['total']}")

    tasks = [run_one(r, mid) for r, mid in targets]
    await asyncio.gather(*tasks, return_exceptions=False)

    wb.save(EXCEL_PATH)
    state.save(STATE_PATH)
    await fetcher.disconnect()
    print(f"\nDone. {counter['done']}/{len(targets)} succeeded, {counter['fail']} failed.")


if __name__ == "__main__":
    asyncio.run(main())
