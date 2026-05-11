# scripts/status.py
"""Print a quick status of the AqmarTofan pipeline: counts, breakdown, and
anything that needs manual review."""
import sys
from pathlib import Path
from collections import Counter

sys.path.insert(0, str(Path(__file__).parent.parent))
from src.state import State
from openpyxl import load_workbook

STATE_PATH = "data/state.json"
EXCEL_PATH = "data/martyrs.xlsx"
MISSING_BIRTH_LOG = "logs/missing_birthdates.log"


def main():
    state = State.load(STATE_PATH)
    print(f"Total processed messages: {len(state.processed_msg_ids)}")
    print(f"Last processed msg_id:     {state.last_processed_msg_id}")

    counts = Counter(state.statuses.values())
    print("\nStatus breakdown:")
    for status, n in counts.most_common():
        print(f"  {status:25} {n}")

    if Path(EXCEL_PATH).exists():
        wb = load_workbook(EXCEL_PATH)
        ws = wb["الشهداء"]
        data_rows = ws.max_row - 2  # 2 header rows
        print(f"\nExcel main rows:      {data_rows}")

        # Coverage of mandatory fields (cols 4=birth, 5=martyrdom, 11=photo)
        birth_filled = 0
        mart_filled = 0
        photo_filled = 0
        for r in range(3, ws.max_row + 1):
            if ws.cell(row=r, column=4).value:
                birth_filled += 1
            if ws.cell(row=r, column=5).value:
                mart_filled += 1
            ph = ws.cell(row=r, column=11).value
            if ph and Path(ph).exists():
                photo_filled += 1
        if data_rows:
            print(f"  birth_date filled:    {birth_filled}/{data_rows} ({birth_filled*100//data_rows}%)")
            print(f"  martyrdom filled:     {mart_filled}/{data_rows} ({mart_filled*100//data_rows}%)")
            print(f"  photo on disk:        {photo_filled}/{data_rows} ({photo_filled*100//data_rows}%)")

        if "النسخ_المكررة" in wb.sheetnames:
            ws2 = wb["النسخ_المكررة"]
            dupes = ws2.max_row - 1  # 1 header row on dupes sheet
            print(f"Duplicate rows:       {dupes}")
    else:
        print(f"\nExcel not yet generated ({EXCEL_PATH} not found).")

    log = Path(MISSING_BIRTH_LOG)
    if log.exists():
        n = len([l for l in log.read_text(encoding="utf-8").splitlines() if l.strip()])
        print(f"\nMissing birthdates queue: {n} msg_ids in {log}")


if __name__ == "__main__":
    main()
