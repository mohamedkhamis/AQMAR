"""Convert data/martyrs.xlsx → data/martyrs.json for the webui.

Run after each pipeline update (or after manual Excel edits). The webui reads
the JSON only — never opens the .xlsx directly.

Uses load_workbook(read_only=True) so it doesn't conflict with the scraper
holding the file open mid-write.
"""
import json
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

DEFAULT_EXCEL = "data/martyrs.xlsx"
DEFAULT_JSON = "data/martyrs.json"

# (json_key, 1-indexed column, type)
FIELDS = [
    ("msg_id",            1,  int),
    ("name",              2,  str),
    ("name_normalized",   3,  str),
    ("birth_date",        4,  str),
    ("martyrdom_date",    5,  str),
    ("city",              6,  str),
    ("military_rank",     7,  str),
    ("weapon",            8,  str),
    ("battalion",         9,  str),
    ("brigade",          10,  str),
    ("photo_path",       11,  str),
    ("posted_date",      13,  str),
    ("message_link",     14,  str),
    ("extraction_status",15,  str),
]


def photo_path_for_webui(p):
    """Rewrite a pipeline-side photo path (e.g. 'data/photos/20.jpg') to
    a relative URL the webui can fetch (../data/photos/20.jpg)."""
    if not p:
        return ""
    p = str(p).replace("\\", "/")
    return "../" + p if not p.startswith("../") else p


def _cell_to(value, caster):
    if value is None or value == "":
        return "" if caster is str else 0
    try:
        return caster(value)
    except (TypeError, ValueError):
        return "" if caster is str else 0


def build_payload(excel_path: str) -> dict:
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb["الشهداء"]
    martyrs = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if row[0] is None:
            continue  # blank trailing row
        record = {}
        for key, col, caster in FIELDS:
            v = row[col - 1] if col - 1 < len(row) else None
            record[key] = _cell_to(v, caster)
        record["photo_path"] = photo_path_for_webui(record["photo_path"])
        martyrs.append(record)
    return {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "channel": "AqmarTofan",
        "martyrs": martyrs,
    }


def main():
    payload = build_payload(DEFAULT_EXCEL)
    Path(DEFAULT_JSON).write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print(f"Wrote {len(payload['martyrs'])} martyrs to {DEFAULT_JSON}")


if __name__ == "__main__":
    main()
