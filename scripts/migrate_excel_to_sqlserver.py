# scripts/migrate_excel_to_sqlserver.py
"""One-shot push of data/martyrs.xlsx → SQL Server `aqmar.dbo.martyrs`.

Run after:
  1. Installing SQL Server (Express or Developer edition) — see README
  2. Running scripts\setup_sqlserver_schema.sql against the aqmar database
  3. SQLSERVER_CONN_STR set in .env (see .env.example)

Idempotent — every row is UPSERTed by msg_id, so re-running is safe.
All migrated rows land with verification_status='unverified' (default),
ready for admin review in the SPA.

Usage:
  .venv\Scripts\activate
  python scripts\migrate_excel_to_sqlserver.py

Optional flags:
  --excel PATH        override the source workbook (default data/martyrs.xlsx)
  --limit N           only migrate the first N rows (handy for smoke tests)
  --chunk N           commit progress every N rows (default 50)
"""
import argparse
import os
import re
import sys
from pathlib import Path

# Force UTF-8 stdout — Arabic chars in row.name crash Windows cp1252.
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

sys.path.insert(0, str(Path(__file__).parent.parent))

from openpyxl import load_workbook
from src.config import load_config
from src.sqlserver_client import make_conn, upsert_martyr, COLUMNS

# Strict YYYY-MM-DD only. SQL Server DATE column rejects anything else (e.g.
# OCR-clipped "29" or year-only "1939"). When a value fails this match, the
# DATE column gets NULL but the raw text is preserved in the ocr_* mirror
# column for the admin to fix during verification.
_ISO_DATE = re.compile(r"^\d{4}-\d{2}-\d{2}$")

# DATETIME2 is more forgiving — accepts "2026-01-07 14:02" and "2026-01-07T14:02:00"
# and any ISO datetime. We still reject obviously broken values.
_ISO_DATETIME = re.compile(r"^\d{4}-\d{2}-\d{2}[ T]\d{2}:\d{2}(:\d{2})?$")


def _sanitize_date(value, *, allow_datetime=False):
    """Return value if it parses as YYYY-MM-DD (or a full DATETIME when
    allow_datetime=True), else None. The original raw string is preserved
    separately in the ocr_* column before this is called."""
    if value is None or value == "":
        return None
    s = str(value).strip()
    if _ISO_DATE.match(s):
        return s
    if allow_datetime and _ISO_DATETIME.match(s):
        return s
    return None

EXCEL_PATH = "data/martyrs.xlsx"
SHEET = "الشهداء"

# (json_key, 1-indexed excel column, type)
# Matches the layout in src/excel_writer.py HEADERS_AR
FIELDS = [
    ("msg_id",            1,  int),
    ("name",              2,  str),
    ("name_normalized",   3,  str),
    ("birth_date",        4,  str),
    ("martyrdom_date",    5,  str),
    ("city",              6,  str),
    ("military_rank",     7,  str),
    ("weapon",            8,  str),
    ("battalion",         9,  str),
    ("brigade",          10,  str),
    ("photo_path",       11,  str),
    ("frame_paths",      12,  str),    # semicolon-separated OCR source frames
    ("posted_date",      13,  str),
    ("message_link",     14,  str),
    ("extraction_status",15,  str),
    ("duplicate_status", 16,  str),
]


def _cell_to(value, caster):
    """Coerce a cell value to the requested type. Empties → None for str/None,
    0 for int/float (matches scripts/excel_to_json.py semantics)."""
    if value is None or value == "":
        return None if caster is str else 0
    # openpyxl returns Excel datetime cells as Python datetime objects; we
    # store the canonical ISO string in birth_date/martyrdom_date/posted_date,
    # so just str() them.
    try:
        return caster(value) if not (caster is str and not isinstance(value, str)) \
            else str(value)
    except (TypeError, ValueError):
        return None if caster is str else 0


def build_payload_from_row(excel_row):
    """Convert one openpyxl row (tuple of cell values) into a dict
    matching src.sqlserver_client.COLUMNS.

    Date fields get a strict YYYY-MM-DD (or datetime) regex check before
    being assigned to the DB DATE / DATETIME2 columns. Values that don't
    match (e.g. OCR clipped to '29' or '1939') get NULL in the structured
    column but the raw text is preserved in the ocr_* mirror column for
    the admin to see and correct during verification.
    """
    payload = {}
    for key, col, caster in FIELDS:
        v = excel_row[col - 1] if col - 1 < len(excel_row) else None
        payload[key] = _cell_to(v, caster)

    # Preserve raw OCR text BEFORE sanitizing the DATE columns.
    raw_birth = payload.get("birth_date")
    raw_martyrdom = payload.get("martyrdom_date")

    # Sanitize date columns — set to NULL when not strict YYYY-MM-DD
    payload["birth_date"] = _sanitize_date(raw_birth)
    payload["martyrdom_date"] = _sanitize_date(raw_martyrdom)
    payload["posted_date"] = _sanitize_date(payload.get("posted_date"), allow_datetime=True)

    # OCR-side mirrors keep the raw (potentially malformed) text so the
    # admin can see what OCR actually extracted and decide what to do.
    payload["ocr_name"] = payload.get("name")
    payload["ocr_birth_date"] = raw_birth
    payload["ocr_martyrdom_date"] = raw_martyrdom

    # Make sure every expected column key exists, even if just as None
    for c in COLUMNS:
        payload.setdefault(c, None)
    return payload


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--excel", default=EXCEL_PATH,
                        help=f"Path to the Excel workbook (default {EXCEL_PATH})")
    parser.add_argument("--limit", type=int, default=None,
                        help="Migrate only the first N rows (smoke test)")
    parser.add_argument("--chunk", type=int, default=50,
                        help="Print progress every N rows (default 50)")
    args = parser.parse_args()

    cfg = load_config()
    if not cfg.sqlserver_conn_str:
        print("ERROR: SQLSERVER_CONN_STR is empty in .env — see .env.example")
        sys.exit(1)

    if not os.path.exists(args.excel):
        print(f"ERROR: Excel workbook not found at {args.excel}")
        sys.exit(1)

    print(f"Reading {args.excel}...")
    wb = load_workbook(args.excel, read_only=True, data_only=True)
    if SHEET not in wb.sheetnames:
        print(f"ERROR: sheet '{SHEET}' not in workbook (have: {wb.sheetnames})")
        sys.exit(1)
    ws = wb[SHEET]

    rows = []
    for excel_row in ws.iter_rows(min_row=3, values_only=True):
        if excel_row[0] is None:    # blank trailing row
            continue
        rows.append(build_payload_from_row(excel_row))
        if args.limit and len(rows) >= args.limit:
            break

    print(f"Loaded {len(rows)} rows from Excel.")
    print(f"Opening SQL Server connection...")
    conn = make_conn(cfg)

    upserted = 0
    failed = 0
    for i, payload in enumerate(rows, 1):
        try:
            upsert_martyr(conn, payload)
            upserted += 1
        except Exception as e:
            print(f"  ERROR upserting msg {payload.get('msg_id')}: {e}")
            failed += 1
        if i % args.chunk == 0:
            print(f"  progress: {i}/{len(rows)} (upserted={upserted}, failed={failed})")

    print(f"\nDone. Upserted {upserted} / {len(rows)} rows ({failed} failed).")

    # Verify final count by querying back
    cur = conn.cursor()
    cur.execute("SELECT COUNT(*) FROM dbo.martyrs")
    total_in_db = cur.fetchone()[0]
    cur.execute("SELECT COUNT(*) FROM dbo.martyrs WHERE verification_status = 'unverified'")
    unverified = cur.fetchone()[0]
    print(f"SQL Server now has {total_in_db} rows ({unverified} unverified, "
          f"{total_in_db - unverified} verified/rejected).")
    conn.close()


if __name__ == "__main__":
    main()
