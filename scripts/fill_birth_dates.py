# scripts/fill_birth_dates.py
"""Overnight Step 2 of the hybrid backfill.

After scripts/phase2_photos_only.py populated 240 rows with
name + martyrdom + photo (but blank birth date), this script:

  1. Scans data/martyrs.xlsx for rows where birth_date is blank.
  2. For each, downloads the full video, extracts frames at sec 28/30/32,
     runs OCR, and updates JUST the birth/martyrdom/rank/city/weapon
     columns of the existing row (does NOT add a new row).
  3. Updates state.json status to "complete" / "partial_*" based on
     what came back.
  4. Saves Excel checkpoint every 5 rows.

Runs sequentially (no concurrency complication) since the goal is
correctness — overnight is fine. Resumable if killed: skips rows that
already have birth_date filled, so re-running just continues where it
left off.

Run in detached PowerShell:
  python scripts\\fill_birth_dates.py
"""
import asyncio
import os
import sys
import logging
import tempfile
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))
from openpyxl import load_workbook
from src.config import load_config
from src.telegram_client import TelegramFetcher
from src.parser_caption import parse_caption
from src.parser_ocr import parse_ocr_text, merge_extractions
from src.frame_extractor import extract_frames
from src.ocr_engine import ocr_image
from src.state import State

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(message)s",
    handlers=[
        logging.FileHandler("logs/fill_birth.log", encoding="utf-8"),
        logging.StreamHandler(),
    ],
)

EXCEL_PATH = "data/martyrs.xlsx"
STATE_PATH = "data/state.json"
FRAMES_DIR = "data/frames"

# Excel column numbers (1-indexed)
COL_MSG_ID = 1
COL_BIRTH = 4
COL_MARTYRDOM = 5
COL_CITY = 6
COL_RANK = 7
COL_WEAPON = 8
COL_FRAMES = 12
COL_STATUS = 15


def determine_status(birth, martyrdom):
    if not birth and not martyrdom:
        return "missing_critical"
    if not birth:
        return "partial_birth"
    if not martyrdom:
        return "partial_martyrdom"
    return "complete"


async def process_msg_video(fetcher, msg_id):
    """Download video, extract frames, OCR, merge — return the fields dict."""
    msg = await fetcher.client.get_messages(fetcher.channel, ids=msg_id)
    if msg is None:
        return None
    # Wrap as TgMessage-shaped for download_video
    from src.telegram_client import TgMessage
    tg = fetcher._to_tg_message(msg)
    if not tg.has_video:
        return None

    frame_paths = []
    with tempfile.TemporaryDirectory() as td:
        video_path = os.path.join(td, f"{msg_id}.mp4")
        await fetcher.download_video(tg, video_path)
        frame_paths = extract_frames(video_path, FRAMES_DIR, msg_id)

    if not frame_paths:
        return None

    extractions = []
    for fp in frame_paths:
        try:
            text = ocr_image(fp)
            extractions.append(parse_ocr_text(text))
        except Exception as e:
            logging.warning(f"OCR failed on {fp}: {e}")

    merged = merge_extractions(extractions) if extractions else None
    return {"fields": merged, "frame_paths": frame_paths}


async def main():
    cfg = load_config()
    fetcher = TelegramFetcher(
        cfg.api_id, cfg.api_hash, cfg.phone, cfg.two_fa_password,
        cfg.session_path, cfg.channel_username,
    )
    await fetcher.connect()

    wb = load_workbook(EXCEL_PATH)
    ws = wb["الشهداء"]
    state = State.load(STATE_PATH)

    # Find rows needing birth date
    targets = []  # (excel_row, msg_id)
    for r in range(3, ws.max_row + 1):
        birth = ws.cell(row=r, column=COL_BIRTH).value
        msg_id = ws.cell(row=r, column=COL_MSG_ID).value
        if msg_id and not birth:
            targets.append((r, int(msg_id)))
    print(f"Found {len(targets)} rows needing birth date.")

    if not targets:
        await fetcher.disconnect()
        return

    done = 0
    for i, (excel_row, msg_id) in enumerate(targets, start=1):
        print(f"[{i}/{len(targets)}] msg {msg_id} — downloading + OCR...")
        try:
            result = await process_msg_video(fetcher, msg_id)
        except Exception as e:
            logging.exception(f"Failed msg {msg_id}: {e}")
            state.mark_processed(msg_id, "video_failed")
            continue

        if result is None or result["fields"] is None:
            print(f"  → no frames extracted, skip")
            state.mark_processed(msg_id, "video_failed")
            continue

        f = result["fields"]
        # Only OVERWRITE if we got something new — never clear existing data
        if f["birth_date"]:
            ws.cell(row=excel_row, column=COL_BIRTH, value=f["birth_date"])
        if f["martyrdom_date"] and not ws.cell(row=excel_row, column=COL_MARTYRDOM).value:
            ws.cell(row=excel_row, column=COL_MARTYRDOM, value=f["martyrdom_date"])
        if f["city"] and not ws.cell(row=excel_row, column=COL_CITY).value:
            ws.cell(row=excel_row, column=COL_CITY, value=f["city"])
        if f["military_rank"] and not ws.cell(row=excel_row, column=COL_RANK).value:
            ws.cell(row=excel_row, column=COL_RANK, value=f["military_rank"])
        if f["weapon"] and not ws.cell(row=excel_row, column=COL_WEAPON).value:
            ws.cell(row=excel_row, column=COL_WEAPON, value=f["weapon"])

        ws.cell(row=excel_row, column=COL_FRAMES, value=";".join(result["frame_paths"]))

        # Refresh status based on current cell values (after the writes)
        birth_now = ws.cell(row=excel_row, column=COL_BIRTH).value
        mart_now = ws.cell(row=excel_row, column=COL_MARTYRDOM).value
        new_status = determine_status(birth_now, mart_now)
        ws.cell(row=excel_row, column=COL_STATUS, value=new_status)
        state.mark_processed(msg_id, new_status)
        done += 1
        print(f"  → {new_status} | birth={birth_now} | mart={mart_now}")

        if done % 5 == 0:
            wb.save(EXCEL_PATH)
            state.save(STATE_PATH)
            print(f"  ✓ checkpoint at {done}/{len(targets)}")

    wb.save(EXCEL_PATH)
    state.save(STATE_PATH)
    await fetcher.disconnect()
    print(f"\nDone. {done}/{len(targets)} birth dates filled in.")


if __name__ == "__main__":
    asyncio.run(main())
