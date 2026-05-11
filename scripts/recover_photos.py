# scripts/recover_photos.py
"""Retry photo downloads for any Excel rows that have an empty photo_path.

Phase 2's first run had a bug where Telegram file references expire
after ~30 min — 95 photos were left as 0-byte files. The download_photo
function now refreshes the reference and rejects 0-byte results, so
running this script after that fix will re-download anything missing.

Workflow:
  1. Delete every 0-byte file under data/photos/
  2. Open data/martyrs.xlsx → find rows with empty photo_path
  3. Rebuild the name -> photo TgMessage index from the channel
  4. Re-fetch each missing photo using the fixed download_photo
  5. Write the new photo_path back into the Excel row, save once at the end

Safe to run multiple times — already-valid photos are left alone.
"""
import asyncio
import os
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))
from openpyxl import load_workbook
from src.config import load_config
from src.telegram_client import TelegramFetcher
from src.parser_caption import parse_caption

EXCEL_PATH = "data/martyrs.xlsx"
PHOTOS_DIR = "data/photos"
PHOTO_COL = 11  # column number of "مسار الصورة" / "Photo Path"
NAME_COL = 2    # column number of "الاسم" / "Name"
MSG_ID_COL = 1


def cleanup_empty_photos() -> int:
    removed = 0
    if not os.path.isdir(PHOTOS_DIR):
        return 0
    for fn in os.listdir(PHOTOS_DIR):
        p = os.path.join(PHOTOS_DIR, fn)
        if os.path.isfile(p) and os.path.getsize(p) == 0:
            os.remove(p)
            removed += 1
    return removed


async def main():
    print(f"Cleaning 0-byte photos from {PHOTOS_DIR}...")
    removed = cleanup_empty_photos()
    print(f"Removed {removed} empty files.")

    print(f"Opening {EXCEL_PATH}...")
    wb = load_workbook(EXCEL_PATH)
    ws = wb["الشهداء"]

    # Identify rows that need a photo
    needs = []  # list of (excel_row, msg_id, name)
    for r in range(3, ws.max_row + 1):
        photo_path = ws.cell(row=r, column=PHOTO_COL).value
        msg_id = ws.cell(row=r, column=MSG_ID_COL).value
        name = ws.cell(row=r, column=NAME_COL).value
        if (not photo_path) or (photo_path and not os.path.exists(photo_path)):
            if msg_id and name:
                needs.append((r, int(msg_id), name))
    print(f"{len(needs)} rows need a photo retry.")
    if not needs:
        print("Nothing to do.")
        return

    cfg = load_config()
    fetcher = TelegramFetcher(
        cfg.api_id, cfg.api_hash, cfg.phone, cfg.two_fa_password,
        cfg.session_path, cfg.channel_username,
    )
    await fetcher.connect()

    print("Building name -> photo index from channel...")
    messages = await fetcher.fetch_all_messages()
    photos = [m for m in messages if m.has_photo]
    name_to_photo = {}
    for p in photos:
        nm = parse_caption(p.caption)["name"]
        if nm and nm not in name_to_photo:
            name_to_photo[nm] = p
    print(f"Indexed {len(name_to_photo)} unique photos by name.")

    recovered = 0
    failed = 0
    for excel_row, msg_id, name in needs:
        paired = name_to_photo.get(name)
        if paired is None:
            print(f"  msg {msg_id}: no paired photo for name {name!r}")
            failed += 1
            continue
        out_path = os.path.join(PHOTOS_DIR, f"{msg_id}.jpg")
        try:
            await fetcher.download_photo(paired, out_path)
            ws.cell(row=excel_row, column=PHOTO_COL, value=out_path)
            print(f"  msg {msg_id}: photo recovered from msg {paired.msg_id}")
            recovered += 1
        except Exception as e:
            print(f"  msg {msg_id}: download failed: {e}")
            failed += 1

        if (recovered + failed) % 20 == 0:
            wb.save(EXCEL_PATH)
            print(f"  ... saved checkpoint at {recovered} recovered / {failed} failed")

    wb.save(EXCEL_PATH)
    await fetcher.disconnect()
    print(f"\nDone. Recovered: {recovered} | Failed: {failed} | Total tried: {len(needs)}")


if __name__ == "__main__":
    asyncio.run(main())
