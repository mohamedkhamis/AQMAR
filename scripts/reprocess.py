# scripts/reprocess.py
"""Re-process a single message and print what the pipeline would extract.

Useful for manually verifying a specific row in the Excel against the
actual frames + photo, or for debugging when one row looks wrong.

Does NOT overwrite the Excel row by default — only prints. Pass
--update to also write the new extraction back into the existing Excel
row (replaces, doesn't append).

Usage:
  python scripts\reprocess.py --msg-id 808
  python scripts\reprocess.py --msg-id 808 --update
"""
import argparse
import asyncio
import sys
import os
from pathlib import Path

# Force UTF-8 on stdout/stderr — Arabic characters in row.name (printed on
# every reprocess) crash Windows console default cp1252 ('charmap codec'
# error). Same fix as scripts/phase3_daily.py.
if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
    sys.stderr.reconfigure(encoding='utf-8', errors='replace')

sys.path.insert(0, str(Path(__file__).parent.parent))
from src.config import load_config
from src.telegram_client import TelegramFetcher
from src.pipeline import process_message
from src.parser_caption import parse_caption
from src.excel_writer import ExcelWriter
from src.state import State
from openpyxl import load_workbook

STATE_PATH = "data/state.json"

EXCEL_PATH = "data/martyrs.xlsx"
PHOTOS_DIR = "data/photos"
FRAMES_DIR = "data/frames"
MISSING_BIRTH_LOG = "logs/missing_birthdates.log"

EXCEL_COLS = {
    "msg_id": 1, "name": 2, "name_normalized": 3, "birth_date": 4,
    "martyrdom_date": 5, "city": 6, "military_rank": 7, "weapon": 8,
    "battalion": 9, "brigade": 10, "photo_path": 11, "frame_paths": 12,
    "posted_date": 13, "message_link": 14, "extraction_status": 15,
    "duplicate_status": 16,
}


def update_excel_row(msg_id: int, row) -> bool:
    if not os.path.exists(EXCEL_PATH):
        return False
    wb = load_workbook(EXCEL_PATH)
    ws = wb["الشهداء"]
    target = None
    for r in range(3, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == msg_id:
            target = r
            break
    if target is None:
        wb.close()
        return False
    for field, col in EXCEL_COLS.items():
        ws.cell(row=target, column=col, value=getattr(row, field))
    wb.save(EXCEL_PATH)
    return True


async def main(msg_id: int, update: bool):
    cfg = load_config()
    fetcher = TelegramFetcher(
        cfg.api_id, cfg.api_hash, cfg.phone, cfg.two_fa_password,
        cfg.session_path, cfg.channel_username,
    )
    await fetcher.connect()

    msg = await fetcher.client.get_messages(cfg.channel_username, ids=msg_id)
    if msg is None:
        print(f"Message {msg_id} not found in @{cfg.channel_username}.")
        await fetcher.disconnect()
        return
    tg = fetcher._to_tg_message(msg)
    if not tg.has_video:
        print(f"Message {msg_id} has no video; nothing to reprocess.")
        await fetcher.disconnect()
        return

    # Find the paired photo by name (looking at adjacent messages)
    name = parse_caption(tg.caption)["name"]
    paired = None
    if name:
        for offset in (-1, -2, -3, 1, 2):
            adj = await fetcher.client.get_messages(cfg.channel_username, ids=msg_id + offset)
            if adj is None:
                continue
            adj_tg = fetcher._to_tg_message(adj)
            if adj_tg.has_photo and parse_caption(adj_tg.caption)["name"] == name:
                paired = adj_tg
                print(f"Paired photo found at msg {paired.msg_id}")
                break

    row = await process_message(
        tg, fetcher, cfg.channel_username,
        PHOTOS_DIR, FRAMES_DIR, MISSING_BIRTH_LOG,
        paired_photo_msg=paired,
    )

    print(f"\nmsg {msg_id}: {row.extraction_status}")
    print(f"  name:        {row.name}")
    print(f"  birth:       {row.birth_date}")
    print(f"  martyrdom:   {row.martyrdom_date}")
    print(f"  city:        {row.city}")
    print(f"  rank:        {row.military_rank}")
    print(f"  weapon:      {row.weapon}")
    print(f"  battalion:   {row.battalion}")
    print(f"  brigade:     {row.brigade}")
    print(f"  photo_path:  {row.photo_path}")
    print(f"  message:     {row.message_link}")

    if update:
        ok = update_excel_row(msg_id, row)
        if ok:
            print(f"\nExcel row for msg {msg_id} updated in {EXCEL_PATH}.")
        else:
            # No existing row → append it via ExcelWriter (handles the workbook
            # init / RTL setup / bold birth+martyrdom styling). Also mark the
            # msg as processed in state.json so the next phase3_daily run skips
            # it. This is the recovery path for messages the daily run failed
            # to append (e.g. msg 874 — cp1252 charmap codec error before fix).
            writer = ExcelWriter(EXCEL_PATH)
            writer.ensure_initialized()
            appended = writer.append_row(row)
            if appended:
                writer.save()
                state = State.load(STATE_PATH)
                state.mark_processed(msg_id, row.extraction_status)
                state.save(STATE_PATH)
                print(f"\nNo existing row; appended msg {msg_id} to {EXCEL_PATH} + marked processed in state.json.")
            else:
                print(f"\nNo existing row, but ExcelWriter.append_row also refused (duplicate msg_id?).")

    await fetcher.disconnect()


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--msg-id", type=int, required=True)
    parser.add_argument("--update", action="store_true",
                        help="Write the new extraction back to the Excel row.")
    args = parser.parse_args()
    asyncio.run(main(args.msg_id, args.update))
