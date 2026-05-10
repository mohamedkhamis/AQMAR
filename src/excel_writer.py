import os
from dataclasses import dataclass, asdict
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

MAIN_SHEET = "الشهداء"
DUPES_SHEET = "النسخ_المكررة"

HEADERS_AR = [
    "المعرف", "الاسم", "الاسم المُطَبَّع", "تاريخ الميلاد", "تاريخ الاستشهاد",
    "المدينة", "الرتبة العسكرية", "السلاح", "الكتيبة", "اللواء",
    "مسار الصورة", "مسار اللقطات", "تاريخ النشر", "رابط الرسالة",
    "حالة الاستخراج", "الحالة",
]
HEADERS_EN = [
    "Msg ID", "Name", "Name Normalized", "Birth Date", "Martyrdom Date",
    "City", "Military Rank", "Weapon", "Battalion", "Brigade",
    "Photo Path", "Frame Paths", "Posted Date", "Message Link",
    "Extraction Status", "Duplicate Status",
]
DUPES_HEADERS_AR = ["المعرف", "الاسم", "السبب", "الدقة", "الحجم", "معرف النسخة المعتمدة", "الرابط"]

@dataclass
class MartyrRow:
    msg_id: int
    name: str
    name_normalized: str
    birth_date: str
    martyrdom_date: str
    city: str
    military_rank: str
    weapon: str
    battalion: str
    brigade: str
    photo_path: str
    frame_paths: str
    posted_date: str
    message_link: str
    extraction_status: str
    duplicate_status: str

@dataclass
class DuplicateRow:
    msg_id: int
    name: str
    reason: str
    resolution: str
    size_mb: float
    kept_msg_id: int
    link: str

class ExcelWriter:
    def __init__(self, path: str):
        self.path = path
        self.wb = None

    def ensure_initialized(self) -> None:
        if os.path.exists(self.path):
            self.wb = load_workbook(self.path)
            return
        self.wb = Workbook()
        ws = self.wb.active
        ws.title = MAIN_SHEET
        ws.sheet_view.rightToLeft = True
        for col, (ar, en) in enumerate(zip(HEADERS_AR, HEADERS_EN), start=1):
            ws.cell(row=1, column=col, value=ar).font = Font(bold=True)
            ws.cell(row=2, column=col, value=en).font = Font(italic=True, size=10)
        ws.freeze_panes = "A3"
        # Highlight birth_date (col 4) and martyrdom_date (col 5)
        ws.cell(row=1, column=4).fill = PatternFill("solid", fgColor="FFF2CC")
        ws.cell(row=1, column=5).fill = PatternFill("solid", fgColor="FCE4D6")
        # Duplicates sheet — single Arabic header row (per spec section 7)
        ws2 = self.wb.create_sheet(DUPES_SHEET)
        ws2.sheet_view.rightToLeft = True
        for col, ar in enumerate(DUPES_HEADERS_AR, start=1):
            ws2.cell(row=1, column=col, value=ar).font = Font(bold=True)
        ws2.freeze_panes = "A2"

    def _existing_msg_ids(self, sheet_name: str) -> set:
        ws = self.wb[sheet_name]
        ids = set()
        for row in ws.iter_rows(min_row=3, max_col=1, values_only=True):
            if row[0] is not None:
                ids.add(int(row[0]))
        return ids

    def append_row(self, row: MartyrRow) -> bool:
        ws = self.wb[MAIN_SHEET]
        if row.msg_id in self._existing_msg_ids(MAIN_SHEET):
            return False
        values = [
            row.msg_id, row.name, row.name_normalized, row.birth_date, row.martyrdom_date,
            row.city, row.military_rank, row.weapon, row.battalion, row.brigade,
            row.photo_path, row.frame_paths, row.posted_date, row.message_link,
            row.extraction_status, row.duplicate_status,
        ]
        ws.append(values)
        # Bold birth + martyrdom date cells
        last = ws.max_row
        ws.cell(row=last, column=4).font = Font(bold=True)
        ws.cell(row=last, column=5).font = Font(bold=True)
        return True

    def append_duplicate(self, row: DuplicateRow) -> None:
        ws = self.wb[DUPES_SHEET]
        ws.append([row.msg_id, row.name, row.reason, row.resolution,
                   row.size_mb, row.kept_msg_id, row.link])

    def save(self) -> None:
        os.makedirs(os.path.dirname(self.path) or ".", exist_ok=True)
        self.wb.save(self.path)
