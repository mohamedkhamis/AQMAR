from openpyxl import load_workbook
from src.excel_writer import ExcelWriter, MartyrRow, DuplicateRow

def test_writes_bilingual_headers(tmp_path):
    p = tmp_path / "out.xlsx"
    w = ExcelWriter(str(p))
    w.ensure_initialized()
    w.save()
    wb = load_workbook(p)
    ws = wb["الشهداء"]
    assert ws["A1"].value == "المعرف"
    assert ws["A2"].value == "Msg ID"
    assert ws["B1"].value == "الاسم"
    assert ws["D1"].value == "تاريخ الميلاد"
    assert ws.sheet_view.rightToLeft is True

def test_appends_row_and_skips_duplicate_msg_id(tmp_path):
    p = tmp_path / "out.xlsx"
    w = ExcelWriter(str(p))
    w.ensure_initialized()
    row = MartyrRow(
        msg_id=808, name="فلان", name_normalized="فلان",
        birth_date="1980-02-12", martyrdom_date="2024-05-17",
        city="غزة", military_rank="نائب قائد سرية", weapon="المدفعية",
        battalion="كتيبة X", brigade="لواء Y",
        photo_path="data/photos/808.jpg", frame_paths="data/frames/808_*.jpg",
        posted_date="2026-05-09 11:50", message_link="https://t.me/AqmarTofan/808",
        extraction_status="complete", duplicate_status="unique",
    )
    assert w.append_row(row) is True   # first time → True
    assert w.append_row(row) is False  # duplicate msg_id → False (skipped)
    w.save()
    wb = load_workbook(p)
    ws = wb["الشهداء"]
    assert ws.max_row == 3  # 2 header rows + 1 data row

def test_writes_duplicates_sheet(tmp_path):
    p = tmp_path / "out.xlsx"
    w = ExcelWriter(str(p))
    w.ensure_initialized()
    w.append_duplicate(DuplicateRow(
        msg_id=807, name="فلان", reason="lower_resolution",
        resolution="854x480", size_mb=12.3, kept_msg_id=808,
        link="https://t.me/AqmarTofan/807",
    ))
    w.save()
    wb = load_workbook(p)
    assert "النسخ_المكررة" in wb.sheetnames
    ws = wb["النسخ_المكررة"]
    assert ws["A2"].value == 807
