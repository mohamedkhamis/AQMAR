"""Tests for the Excel → JSON converter used by the webui."""
import json
import sys
from pathlib import Path
from openpyxl import Workbook

sys.path.insert(0, str(Path(__file__).parent.parent))

from scripts.excel_to_json import build_payload, photo_path_for_webui


def _make_xlsx(tmp_path, rows):
    """Create a tiny xlsx file shaped like data/martyrs.xlsx for testing."""
    wb = Workbook()
    ws = wb.active
    ws.title = "الشهداء"
    # Two header rows (the real Excel uses Arabic + English headers)
    for col, name in enumerate(
        ["المعرف", "الاسم", "الاسم المُطَبَّع", "تاريخ الميلاد",
         "تاريخ الاستشهاد", "المدينة", "الرتبة العسكرية", "السلاح",
         "الكتيبة", "اللواء", "مسار الصورة", "مسار اللقطات",
         "تاريخ النشر", "رابط الرسالة", "حالة الاستخراج", "الحالة"],
        start=1,
    ):
        ws.cell(row=1, column=col, value=name)
        ws.cell(row=2, column=col, value=f"en_{col}")
    for r_offset, row in enumerate(rows, start=3):
        for c_offset, v in enumerate(row, start=1):
            ws.cell(row=r_offset, column=c_offset, value=v)
    path = tmp_path / "martyrs.xlsx"
    wb.save(path)
    return path


def test_photo_path_for_webui_rewrites_to_relative():
    """Photo path in Excel is `data/photos/20.jpg`; webui needs `../data/photos/20.jpg`."""
    assert photo_path_for_webui("data/photos/20.jpg") == "../data/photos/20.jpg"
    assert photo_path_for_webui("data\\photos\\20.jpg") == "../data/photos/20.jpg"
    assert photo_path_for_webui("") == ""
    assert photo_path_for_webui(None) == ""


def test_build_payload_extracts_one_row(tmp_path):
    xlsx = _make_xlsx(
        tmp_path,
        [[20, "مهدي جبر كوارع", "مهدي جبر كوارع", "1977-12-24", "2025-05-15",
          "", "قائد كتيبة", "", "كتيبة س", "لواء رفح",
          "data/photos/20.jpg", "data/frames/20_*.jpg",
          "2024-05-08 14:32", "https://t.me/AqmarTofan/20", "complete", "unique"]],
    )
    payload = build_payload(str(xlsx))
    assert payload["channel"] == "AqmarTofan"
    assert "generated_at" in payload
    assert len(payload["martyrs"]) == 1
    m = payload["martyrs"][0]
    assert m["msg_id"] == 20
    assert m["name"] == "مهدي جبر كوارع"
    assert m["birth_date"] == "1977-12-24"
    assert m["photo_path"] == "../data/photos/20.jpg"
    assert m["message_link"] == "https://t.me/AqmarTofan/20"
    assert m["extraction_status"] == "complete"


def test_build_payload_empty_cells_become_empty_string(tmp_path):
    """None / blank cells must serialize as '' not null for clean Alpine bindings."""
    xlsx = _make_xlsx(
        tmp_path,
        [[42, "اسم", "اسم", "", "", "", "", "", "", "", "", "", "", "", "", ""]],
    )
    payload = build_payload(str(xlsx))
    m = payload["martyrs"][0]
    for key in ("birth_date", "martyrdom_date", "city", "military_rank",
                "weapon", "battalion", "brigade", "photo_path",
                "posted_date", "message_link", "extraction_status"):
        assert m[key] == "", f"expected empty string for {key}, got {m[key]!r}"


def test_build_payload_multiple_rows_preserves_order(tmp_path):
    xlsx = _make_xlsx(
        tmp_path,
        [[1, "أ", "", "", "", "", "", "", "", "", "", "", "", "", "", ""],
         [2, "ب", "", "", "", "", "", "", "", "", "", "", "", "", "", ""],
         [3, "ج", "", "", "", "", "", "", "", "", "", "", "", "", "", ""]],
    )
    payload = build_payload(str(xlsx))
    assert [m["msg_id"] for m in payload["martyrs"]] == [1, 2, 3]
